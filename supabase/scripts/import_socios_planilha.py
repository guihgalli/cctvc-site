#!/usr/bin/env python3
"""Gera SQL de importação a partir de Relatorio_Associados_e_Dependentes.xlsx."""
from __future__ import annotations

import json
import random
import secrets
import sys
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Instale openpyxl: pip install openpyxl", file=sys.stderr)
    raise

ROOT = Path(__file__).resolve().parents[2]
XLSX = ROOT / "Relatorio_Associados_e_Dependentes.xlsx"
OUT_SQL = Path(__file__).resolve().parent / "import_socios_planilha.sql"
OUT_ADMIN = Path(__file__).resolve().parent / "import_admin_credentials.json"

# Colunas aba Titulares (export wide)
COL_NOME, COL_MAT, COL_USER, COL_SENHA = 0, 1, 2, 3
COL_ADM_MAT, COL_ADM_NOME, COL_ADMISSAO = 5, 6, 7
COL_CT_MAT, COL_CT_NOME = 9, 10
COL_TEL_RES, COL_CEL, COL_CPF = 11, 12, 13
COL_EMAIL_COBR, COL_EMAIL_CORR = 14, 15
MIN_COLS = 16

# Aba Dependentes (login + bloco de contato opcional após nascimento)
DEP_MAT, DEP_NOME, DEP_USER, DEP_SENHA = 0, 6, 7, 8
DEP_PARENTESCO, DEP_SEXO, DEP_NASC = 9, 10, 11
DEP_TEL_RES, DEP_CEL, DEP_CPF = 12, 13, 14
DEP_EMAIL_COBR, DEP_EMAIL_CORR = 15, 16
MIN_DEP_COLS = 17


def sql_str(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def sql_date(value: date | None) -> str:
    if value is None:
        return "NULL"
    return sql_str(value.isoformat())


def sql_int(value: int | None) -> str:
    if value is None:
        return "NULL"
    return str(int(value))


def excel_date(value: object) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    return None


def cpf_from_senha(senha: str) -> str:
    return (senha.zfill(6) + "00000")[:11]


def pad_row(row: tuple, size: int = MIN_COLS) -> tuple:
    row = tuple(row)
    return row + (None,) * (size - len(row))


def clean_cpf(value: object) -> str | None:
    if value is None:
        return None
    import re

    digits = re.sub(r"\D", "", str(value))
    return digits if len(digits) == 11 else None


def clean_phone(value: object) -> str | None:
    if value is None:
        return None
    import re

    digits = re.sub(r"\D", "", str(value))
    return digits if len(digits) in (10, 11) else None


def clean_email(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip().lower()
    return text if "@" in text else None


def contact_from_row(tel_res: object, cel: object, cpf: object, email_cobr: object, email_corr: object) -> dict:
    return {
        "cpf": clean_cpf(cpf),
        "telefone": clean_phone(cel) or clean_phone(tel_res),
        "email": clean_email(email_cobr) or clean_email(email_corr),
    }


def resolve_contact(
    *,
    matricula: int | None,
    codigo: str,
    nome: str,
    senha: str,
    inline: dict,
    contacts_by_mat: dict[int, dict],
    contacts_by_nome: dict[str, dict],
    for_dependente: bool = False,
) -> tuple[str, str, str]:
    contact: dict = {}
    if any(inline.values()):
        contact = inline
    elif for_dependente:
        contact = contacts_by_nome.get(nome.upper(), {})
    elif matricula is not None and matricula in contacts_by_mat:
        candidate = contacts_by_mat[matricula]
        if candidate.get("nome", "").upper() == nome.upper():
            contact = candidate
    if not contact and not for_dependente:
        contact = contacts_by_nome.get(nome.upper(), {})

    cpf = contact.get("cpf") or cpf_from_senha(senha)
    return cpf, contact.get("email") or "", contact.get("telefone") or ""


def codigo_titular(codigo: str) -> str:
    return codigo[:3] + "0"


def load_rows() -> tuple[list[dict], list[dict], dict[int, dict], dict[int, dict], dict[str, dict]]:
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

    titulares: list[dict] = []
    contacts_by_mat: dict[int, dict] = {}
    contacts_by_nome: dict[str, dict] = {}
    admissao_by_mat: dict[int, date | None] = {}

    for raw in wb["Titulares"].iter_rows(min_row=2, values_only=True):
        row = pad_row(raw)

        if row[COL_NOME] and row[COL_USER]:
            titulares.append(
                {
                    "nome": str(row[COL_NOME]).strip(),
                    "matricula": int(row[COL_MAT]) if row[COL_MAT] is not None else None,
                    "user": str(row[COL_USER]).zfill(4),
                    "senha": str(row[COL_SENHA]).zfill(6),
                }
            )

        if row[COL_ADM_MAT] is not None and row[COL_ADM_NOME]:
            admissao_by_mat[int(row[COL_ADM_MAT])] = excel_date(row[COL_ADMISSAO])

        if row[COL_CT_MAT] is not None and row[COL_CT_NOME]:
            mat = int(row[COL_CT_MAT])
            nome = str(row[COL_CT_NOME]).strip()
            contact = contact_from_row(row[COL_TEL_RES], row[COL_CEL], row[COL_CPF], row[COL_EMAIL_COBR], row[COL_EMAIL_CORR])
            contact["nome"] = nome
            contacts_by_mat[mat] = contact
            contacts_by_nome[nome.upper()] = contact

    titular_meta: dict[int, dict] = {}
    dependentes: list[dict] = []
    for raw in wb["Dependentes"].iter_rows(min_row=2, values_only=True):
        row = pad_row(raw, MIN_DEP_COLS)
        if not row[DEP_NOME] or not row[DEP_USER]:
            continue
        matricula = int(row[DEP_MAT]) if row[DEP_MAT] is not None else None
        if matricula is not None and matricula not in titular_meta:
            titular_meta[matricula] = {
                "categoria_clube": str(row[2]).strip() if row[2] else None,
                "data_nascimento": excel_date(row[3]),
                "data_admissao": admissao_by_mat.get(matricula) or excel_date(row[4]),
            }
        dependentes.append(
            {
                "matricula": matricula,
                "nome": str(row[DEP_NOME]).strip(),
                "user": str(row[DEP_USER]).zfill(4),
                "senha": str(row[DEP_SENHA]).zfill(6),
                "numero_dependente": int(row[5]) if row[5] is not None else None,
                "parentesco": str(row[DEP_PARENTESCO]).strip() if row[DEP_PARENTESCO] else None,
                "sexo": str(row[DEP_SEXO]).strip().upper()[:1] if row[DEP_SEXO] else None,
                "data_nascimento": excel_date(row[DEP_NASC]),
                "inline_contact": contact_from_row(
                    row[DEP_TEL_RES], row[DEP_CEL], row[DEP_CPF], row[DEP_EMAIL_COBR], row[DEP_EMAIL_CORR]
                ),
            }
        )

    wb.close()
    return titulares, dependentes, titular_meta, contacts_by_mat, contacts_by_nome


def pick_admin_code(used: set[str]) -> str:
    rng = random.SystemRandom()
    for _ in range(10_000):
        code = f"{rng.randint(0, 9999):04d}"
        if code not in used:
            return code
    raise RuntimeError("Não foi possível gerar código admin livre")


def load_admin_credentials(used: set[str]) -> tuple[str, str]:
    if OUT_ADMIN.exists():
        try:
            data = json.loads(OUT_ADMIN.read_text(encoding="utf-8"))
            code = str(data.get("codigo_usuario", "")).zfill(4)
            senha = str(data.get("senha", "")).zfill(6)
            if code and senha and code not in used:
                return code, senha
        except (json.JSONDecodeError, OSError):
            pass
    return pick_admin_code(used), f"{secrets.randbelow(1_000_000):06d}"


def main() -> None:
    titulares, dependentes, titular_meta, contacts_by_mat, contacts_by_nome = load_rows()
    used_codes = {t["user"] for t in titulares} | {d["user"] for d in dependentes}

    admin_code, admin_senha = load_admin_credentials(used_codes)
    admin_cpf = cpf_from_senha(admin_senha)
    admin_nome = "Administrador CCTVC"

    mat_to_user: dict[int, str] = {}
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    for row in wb["Titulares"].iter_rows(min_row=2, values_only=True):
        if row[0] and row[1] is not None and row[2]:
            mat_to_user[int(row[1])] = str(row[2]).zfill(4)
    wb.close()

    insert_cols = (
        "codigo_usuario, cpf, nome, email, telefone, perfil, tipo_socio, categoria_socio, "
        "titular_id, matricula, categoria_clube, data_nascimento, data_admissao, "
        "parentesco, sexo, numero_dependente, senha_hash"
    )

    lines: list[str] = [
        "-- Importação: Relatorio_Associados_e_Dependentes.xlsx (todos os campos)",
        "SET search_path = public, extensions;",
        "BEGIN;",
        "",
        "DELETE FROM sessoes;",
        "DELETE FROM reserva_participantes;",
        "DELETE FROM reservas;",
        "DELETE FROM liberacoes_quadra_locacao;",
        "DELETE FROM usuarios;",
        "",
        f"INSERT INTO usuarios ({insert_cols})",
        f"VALUES ({sql_str(admin_code)}, {sql_str(admin_cpf)}, {sql_str(admin_nome)}, '', '', 'admin', 'socio', NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, crypt({sql_str(admin_senha)}, gen_salt('bf')));",
        "",
    ]

    for t in titulares:
        meta = titular_meta.get(t["matricula"] or -1, {})
        cpf, email, telefone = resolve_contact(
            matricula=t["matricula"],
            codigo=t["user"],
            nome=t["nome"],
            senha=t["senha"],
            inline={},
            contacts_by_mat=contacts_by_mat,
            contacts_by_nome=contacts_by_nome,
            for_dependente=False,
        )
        cat = meta.get("categoria_clube")
        lines.append(
            f"INSERT INTO usuarios ({insert_cols})"
            f" VALUES ({sql_str(t['user'])}, {sql_str(cpf)}, {sql_str(t['nome'])}, {sql_str(email)}, {sql_str(telefone)}, 'usuario', 'socio', 'titular', NULL,"
            f" {sql_int(t['matricula'])}, {sql_str(cat) if cat else 'NULL'}, {sql_date(meta.get('data_nascimento'))}, {sql_date(meta.get('data_admissao'))},"
            f" NULL, NULL, NULL, crypt({sql_str(t['senha'])}, gen_salt('bf')));"
        )

    lines.append("")

    for d in dependentes:
        cpf, email, telefone = resolve_contact(
            matricula=d["matricula"],
            codigo=d["user"],
            nome=d["nome"],
            senha=d["senha"],
            inline=d.get("inline_contact", {}),
            contacts_by_mat=contacts_by_mat,
            contacts_by_nome=contacts_by_nome,
            for_dependente=True,
        )
        titular_code = mat_to_user.get(d["matricula"] or -1) or codigo_titular(d["user"])
        parentesco = d.get("parentesco")
        sexo = d.get("sexo")
        lines.append(
            f"INSERT INTO usuarios ({insert_cols})"
            f" VALUES ({sql_str(d['user'])}, {sql_str(cpf)}, {sql_str(d['nome'])}, {sql_str(email)}, {sql_str(telefone)}, 'usuario', 'socio', 'dependente',"
            f" (SELECT id FROM usuarios WHERE codigo_usuario = {sql_str(titular_code)} LIMIT 1),"
            f" {sql_int(d['matricula'])}, NULL, {sql_date(d.get('data_nascimento'))}, NULL,"
            f" {sql_str(parentesco) if parentesco else 'NULL'}, {sql_str(sexo) if sexo else 'NULL'}, {sql_int(d.get('numero_dependente'))},"
            f" crypt({sql_str(d['senha'])}, gen_salt('bf')));"
        )

    lines.extend(["", "COMMIT;", ""])

    OUT_SQL.write_text("\n".join(lines), encoding="utf-8")
    OUT_ADMIN.write_text(
        json.dumps(
            {
                "codigo_usuario": admin_code,
                "senha": admin_senha,
                "nome": admin_nome,
                "titulares": len(titulares),
                "dependentes": len(dependentes),
            },
            indent=2,
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    print(f"Titulares: {len(titulares)}")
    print(f"Dependentes: {len(dependentes)}")
    print(f"Admin: {admin_code} / {admin_senha}")
    print(f"SQL: {OUT_SQL}")


if __name__ == "__main__":
    main()
