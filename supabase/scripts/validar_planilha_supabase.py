#!/usr/bin/env python3
"""Compara Relatorio_Associados_e_Dependentes.xlsx com usuários no Supabase (via export SQL ou REST)."""
from __future__ import annotations

import json
import os
import re
import sys
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Instale openpyxl: pip install openpyxl", file=sys.stderr)
    raise

ROOT = Path(__file__).resolve().parents[2]
XLSX = ROOT / "Relatorio_Associados_e_Dependentes.xlsx"


def excel_date(value: object) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(text[:10], fmt).date().isoformat()
        except ValueError:
            continue
    return None


def cpf_from_senha(senha: str) -> str:
    return (str(senha).zfill(6) + "00000")[:11]


def norm_str(v: object) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def load_planilha() -> dict[str, dict]:
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    users: dict[str, dict] = {}

    titular_meta: dict[int, dict] = {}
    for row in wb["Titulares"].iter_rows(min_row=2, values_only=True):
        nome, matricula, user, senha = row[0], row[1], row[2], row[3]
        if not nome or not user:
            continue
        code = str(user).zfill(4)
        users[code] = {
            "codigo_usuario": code,
            "nome": str(nome).strip().upper(),
            "matricula": int(matricula) if matricula is not None else None,
            "senha": str(senha).zfill(6),
            "cpf_esperado": cpf_from_senha(str(senha)),
            "categoria_socio": "titular",
            "categoria_clube": None,
            "data_nascimento": None,
            "data_admissao": None,
            "parentesco": None,
            "sexo": None,
            "numero_dependente": None,
            "email_planilha": None,
            "telefone_planilha": None,
            "cpf_planilha": None,
        }

    for row in wb["Dependentes"].iter_rows(min_row=2, values_only=True):
        if not row[6] or not row[7]:
            continue
        matricula = int(row[0]) if row[0] is not None else None
        if matricula is not None and matricula not in titular_meta:
            titular_meta[matricula] = {
                "categoria_clube": norm_str(row[2]),
                "data_nascimento": excel_date(row[3]),
                "data_admissao": excel_date(row[4]),
            }
        code = str(row[7]).zfill(4)
        users[code] = {
            "codigo_usuario": code,
            "nome": str(row[6]).strip().upper(),
            "matricula": matricula,
            "senha": str(row[8]).zfill(6),
            "cpf_esperado": cpf_from_senha(str(row[8])),
            "categoria_socio": "dependente",
            "categoria_clube": None,
            "data_nascimento": excel_date(row[11]),
            "data_admissao": None,
            "parentesco": norm_str(row[9]),
            "sexo": norm_str(row[10]),
            "numero_dependente": int(row[5]) if row[5] is not None else None,
            "email_planilha": None,
            "telefone_planilha": None,
            "cpf_planilha": None,
        }

    wb.close()

    for code, u in users.items():
        if u["categoria_socio"] == "titular" and u["matricula"] in titular_meta:
            meta = titular_meta[u["matricula"]]
            u["categoria_clube"] = meta.get("categoria_clube")
            u["data_nascimento"] = meta.get("data_nascimento")
            u["data_admissao"] = meta.get("data_admissao")

    return users


def load_supabase_json(path: Path) -> dict[str, dict]:
    data = json.loads(path.read_text(encoding="utf-8"))
    out: dict[str, dict] = {}
    for u in data:
        code = str(u.get("codigo_usuario", "")).zfill(4)
        if not code:
            continue
        out[code] = {
            "codigo_usuario": code,
            "nome": (u.get("nome") or "").strip().upper(),
            "matricula": u.get("matricula"),
            "cpf": (u.get("cpf") or "").strip(),
            "email": (u.get("email") or "").strip(),
            "telefone": (u.get("telefone") or "").strip(),
            "categoria_socio": u.get("categoria_socio"),
            "categoria_clube": u.get("categoria_clube"),
            "data_nascimento": u.get("data_nascimento"),
            "data_admissao": u.get("data_admissao"),
            "parentesco": u.get("parentesco"),
            "sexo": u.get("sexo"),
            "numero_dependente": u.get("numero_dependente"),
            "tipo_socio": u.get("tipo_socio"),
            "perfil": u.get("perfil"),
        }
    return out


def fetch_supabase_via_sql_export() -> dict[str, dict]:
    """Use arquivo gerado manualmente ou variável SUPABASE_USERS_JSON."""
    env_path = os.environ.get("SUPABASE_USERS_JSON")
    if env_path and Path(env_path).exists():
        return load_supabase_json(Path(env_path))
    default = Path(__file__).resolve().parent / "supabase_usuarios_snapshot.json"
    if default.exists():
        return load_supabase_json(default)
    return {}


def main() -> None:
    planilha = load_planilha()
    db = fetch_supabase_via_sql_export()

    print("=== Planilha xlsx ===")
    print(f"Usuários na planilha: {len(planilha)}")
    titulares = sum(1 for u in planilha.values() if u["categoria_socio"] == "titular")
    deps = sum(1 for u in planilha.values() if u["categoria_socio"] == "dependente")
    print(f"  Titulares: {titulares}, Dependentes: {deps}")
    print("Colunas disponíveis: nome, matrícula, usuário, senha (+ campos planilha dependentes)")
    print("  email, telefone e CPF real: NÃO existem na planilha")
    print()

    if not db:
        print("Snapshot Supabase não encontrado.")
        print("Exporte com:")
        print("  SELECT row_to_json(u) FROM usuarios u WHERE perfil <> 'admin';")
        print("Salve em supabase/scripts/supabase_usuarios_snapshot.json")
        return

    socios_db = {k: v for k, v in db.items() if v.get("tipo_socio") == "socio" and v.get("perfil") != "admin"}
    print("=== Supabase ===")
    print(f"Sócios no banco: {len(socios_db)}")
    com_email = sum(1 for u in socios_db.values() if u["email"])
    com_tel = sum(1 for u in socios_db.values() if u["telefone"])
    com_cpf_sint = sum(1 for u in socios_db.values() if u["cpf"].endswith("00000"))
    print(f"  Com e-mail: {com_email}")
    print(f"  Com telefone: {com_tel}")
    print(f"  CPF sintético (senha+00000): {com_cpf_sint}")
    print()

    plan_codes = set(planilha)
    db_codes = set(socios_db)
    print("=== Cobertura ===")
    print(f"Só na planilha: {len(plan_codes - db_codes)}")
    print(f"Só no Supabase: {len(db_codes - plan_codes)}")
    print(f"Em ambos: {len(plan_codes & db_codes)}")
    print()

    nome_diff = []
    cpf_diff = []
    mat_diff = []
    for code in sorted(plan_codes & db_codes):
        p, d = planilha[code], socios_db[code]
        if p["nome"] != d["nome"]:
            nome_diff.append((code, p["nome"], d["nome"]))
        if p["cpf_esperado"] != d["cpf"]:
            cpf_diff.append((code, p["cpf_esperado"], d["cpf"]))
        if p["matricula"] != d["matricula"]:
            mat_diff.append((code, p["matricula"], d["matricula"]))

    print(f"Nomes divergentes: {len(nome_diff)}")
    for item in nome_diff[:5]:
        print(f"  {item[0]}: planilha={item[1]} | db={item[2]}")
    print(f"CPF divergente do esperado (import): {len(cpf_diff)}")
    print(f"Matrículas divergentes: {len(mat_diff)}")
    print()
    print("=== Conclusão e-mail / telefone / CPF ===")
    print("A planilha NÃO contém e-mail, telefone nem CPF cadastral.")
    print("O import gera CPF placeholder a partir da senha (6 dígitos + 00000).")
    print("Campos vazios na UI são esperados até cadastro manual ou nova planilha.")


if __name__ == "__main__":
    main()
