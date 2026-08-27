import re
from pathlib import Path

import openpyxl


def extract_users_from_sql(sql_path: Path) -> dict[str, dict]:
    text = sql_path.read_text(encoding="utf-8") if sql_path.exists() else ""
    users: dict[str, dict] = {}
    for code, cpf, nome in re.findall(
        r"VALUES \('(\d{4})', '(\d{11})', '((?:''|[^'])*)'", text
    ):
        users[code] = {"cpf": cpf, "nome": nome.replace("''", "'")}
    return users


def load_xlsx_users() -> dict[str, dict]:
    wb = openpyxl.load_workbook(
        Path(__file__).resolve().parents[2] / "Relatorio_Associados_e_Dependentes.xlsx",
        read_only=True,
        data_only=True,
    )
    users: dict[str, dict] = {}
    for row in wb["Titulares"].iter_rows(min_row=2, values_only=True):
        if row[0] and row[2]:
            code = str(row[2]).zfill(4)
            users[code] = {
                "nome": str(row[0]).strip(),
                "senha": str(row[3]).zfill(6),
            }
    for row in wb["Dependentes"].iter_rows(min_row=2, values_only=True):
        if row[6] and row[7]:
            code = str(row[7]).zfill(4)
            users[code] = {
                "nome": str(row[6]).strip(),
                "senha": str(row[8]).zfill(6),
            }
    wb.close()
    return users


def main() -> None:
    base = Path(__file__).resolve().parent
    old = extract_users_from_sql(base / "import_socios_planilha.sql")
    new = load_xlsx_users()
    print("old_sql", len(old), "new_xlsx", len(new))
    added = sorted(set(new) - set(old))
    removed = sorted(set(old) - set(new))
    name_changes = [
        code
        for code in set(new) & set(old)
        if new[code]["nome"] != old[code]["nome"]
    ]
    senha_changes = [
        code
        for code in set(new) & set(old)
        if new[code].get("senha") and old[code]["cpf"][:6] != new[code]["senha"]
    ]
    print("added", len(added), added[:10])
    print("removed", len(removed), removed[:10])
    print("name_changes", len(name_changes))
    for code in name_changes[:10]:
        print(f"  {code}: {old[code]['nome']} -> {new[code]['nome']}")
    print("senha_changes", len(senha_changes), senha_changes[:10])


if __name__ == "__main__":
    main()
