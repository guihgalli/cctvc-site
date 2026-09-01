#!/usr/bin/env python3
"""Gera SQL de UPDATE (cpf, email, telefone) a partir da planilha com dados de contato."""
from __future__ import annotations

import re
import sys
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Instale openpyxl: pip install openpyxl", file=sys.stderr)
    raise

ROOT = Path(__file__).resolve().parents[2]
DEFAULT_XLSX = ROOT / "Relatorio_Associados_e_Dependentes.xlsx"
OUT_SQL = Path(__file__).resolve().parent / "026_sync_contatos_planilha.sql"

# Aba Titulares (layout wide)
COL_NOME = 0
COL_MATRICULA = 1
COL_USUARIO = 2
COL_SENHA = 3
COL_ADM_MAT = 5
COL_ADM_NOME = 6
COL_ADMISSAO = 7
COL_CT_MAT = 9
COL_CT_NOME = 10
COL_TEL_RES = 11
COL_CEL = 12
COL_CPF = 13
COL_EMAIL_COBR = 14
COL_EMAIL_CORR = 15
MIN_TIT_COLS = 16

# Aba Dependentes
DEP_MAT = 0
DEP_NOME = 6
DEP_USUARIO = 7
DEP_SENHA = 8
DEP_PARENTESCO = 9
DEP_SEXO = 10
DEP_NASC = 11
DEP_TEL_RES = 12
DEP_CEL = 13
DEP_CPF = 14
DEP_EMAIL_COBR = 15
DEP_EMAIL_CORR = 16
MIN_DEP_COLS = 17


def sql_str(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def pad_row(row: tuple, size: int) -> tuple:
    row = tuple(row)
    return row + (None,) * (size - len(row))


def excel_date(value: object) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    return None


def clean_cpf(value: object) -> str | None:
    if value is None:
        return None
    digits = re.sub(r"\D", "", str(value))
    return digits if len(digits) == 11 else None


def clean_phone(value: object) -> str | None:
    if value is None:
        return None
    digits = re.sub(r"\D", "", str(value))
    return digits if len(digits) in (10, 11) else None


def clean_email(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip().lower()
    return text if "@" in text else None


def cpf_from_senha(senha: str) -> str:
    return (senha.zfill(6) + "00000")[:11]


def contact_from_row(
    tel_res: object,
    cel: object,
    cpf: object,
    email_cobr: object,
    email_corr: object,
) -> dict:
    return {
        "cpf": clean_cpf(cpf),
        "telefone": clean_phone(cel) or clean_phone(tel_res),
        "email": clean_email(email_cobr) or clean_email(email_corr),
    }


def load_planilha(xlsx: Path) -> tuple[dict[str, dict], dict[str, dict], dict[int, object], dict[str, dict]]:
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws_tit = wb["Titulares"]
    ws_dep = wb["Dependentes"]

    titulares: dict[str, dict] = {}
    contacts_by_mat: dict[int, dict] = {}
    contacts_by_nome: dict[str, dict] = {}
    admissao: dict[int, object] = {}

    for raw in ws_tit.iter_rows(min_row=2, values_only=True):
        row = pad_row(raw, MIN_TIT_COLS)

        if row[COL_NOME] and row[COL_USUARIO]:
            user = str(row[COL_USUARIO]).zfill(4)
            matricula = int(row[COL_MATRICULA]) if row[COL_MATRICULA] is not None else None
            senha = str(row[COL_SENHA]).zfill(6) if row[COL_SENHA] else None
            titulares[user] = {
                "codigo_usuario": user,
                "nome": str(row[COL_NOME]).strip(),
                "matricula": matricula,
                "senha": senha,
            }

        if row[COL_ADM_MAT] is not None and row[COL_ADM_NOME]:
            admissao[int(row[COL_ADM_MAT])] = row[COL_ADMISSAO]

        if row[COL_CT_MAT] is not None and row[COL_CT_NOME]:
            mat = int(row[COL_CT_MAT])
            nome = str(row[COL_CT_NOME]).strip()
            contact = contact_from_row(row[COL_TEL_RES], row[COL_CEL], row[COL_CPF], row[COL_EMAIL_COBR], row[COL_EMAIL_CORR])
            contact["nome"] = nome
            contacts_by_mat[mat] = contact
            contacts_by_nome[nome.upper()] = contact

    dependentes: dict[str, dict] = {}
    for raw in ws_dep.iter_rows(min_row=2, values_only=True):
        row = pad_row(raw, MIN_DEP_COLS)
        if not row[DEP_NOME] or not row[DEP_USUARIO]:
            continue

        user = str(row[DEP_USUARIO]).zfill(4)
        senha = str(row[DEP_SENHA]).zfill(6) if row[DEP_SENHA] else None
        nome = str(row[DEP_NOME]).strip()
        inline = contact_from_row(
            row[DEP_TEL_RES], row[DEP_CEL], row[DEP_CPF], row[DEP_EMAIL_COBR], row[DEP_EMAIL_CORR]
        )

        dependentes[user] = {
            "codigo_usuario": user,
            "nome": nome,
            "matricula": int(row[DEP_MAT]) if row[DEP_MAT] is not None else None,
            "senha": senha,
            "parentesco": str(row[DEP_PARENTESCO]).strip() if row[DEP_PARENTESCO] else None,
            "sexo": str(row[DEP_SEXO]).strip().upper()[:1] if row[DEP_SEXO] else None,
            "numero_dependente": int(row[5]) if row[5] is not None else None,
            "data_nascimento": excel_date(row[DEP_NASC]),
            "inline_contact": inline,
        }

    wb.close()
    return titulares, dependentes, admissao, {
        "by_mat": contacts_by_mat,
        "by_nome": contacts_by_nome,
    }


def resolve_contact(
    *,
    matricula: int | None,
    codigo_usuario: str,
    nome: str,
    senha: str | None,
    inline: dict,
    contacts: dict[str, dict],
    for_dependente: bool = False,
) -> tuple[str | None, str, str, bool]:
    by_mat = contacts["by_mat"]
    by_nome = contacts["by_nome"]

    contact = {}
    if any(inline.values()):
        contact = inline
    elif for_dependente:
        contact = by_nome.get(nome.upper(), {})
    elif matricula is not None and matricula in by_mat:
        candidate = by_mat[matricula]
        if candidate.get("nome", "").upper() == nome.upper():
            contact = candidate
    if not contact and not for_dependente:
        contact = by_nome.get(nome.upper(), {})

    cpf = contact.get("cpf") or (cpf_from_senha(senha) if senha else None)
    email = contact.get("email") or ""
    telefone = contact.get("telefone") or ""
    cpf_real = bool(contact.get("cpf"))
    return cpf, email, telefone, cpf_real


def build_updates(xlsx: Path = DEFAULT_XLSX) -> tuple[list[dict], list[dict]]:
    titulares, dependentes, admissao, contacts = load_planilha(xlsx)
    tit_updates: list[dict] = []
    dep_updates: list[dict] = []

    for user, titular in sorted(titulares.items()):
        mat = titular["matricula"]
        cpf, email, telefone, cpf_real = resolve_contact(
            matricula=mat,
            codigo_usuario=user,
            nome=titular["nome"],
            senha=titular["senha"],
            inline={},
            contacts=contacts,
            for_dependente=False,
        )
        data_admissao = excel_date(admissao.get(mat)) if mat is not None else None

        tit_updates.append(
            {
                "codigo_usuario": user,
                "nome": titular["nome"],
                "cpf": cpf,
                "email": email,
                "telefone": telefone,
                "data_admissao": data_admissao.isoformat() if data_admissao else None,
                "cpf_real": cpf_real,
            }
        )

    for user, dep in sorted(dependentes.items()):
        cpf, email, telefone, cpf_real = resolve_contact(
            matricula=dep["matricula"],
            codigo_usuario=user,
            nome=dep["nome"],
            senha=dep["senha"],
            inline=dep["inline_contact"],
            contacts=contacts,
            for_dependente=True,
        )
        dep_updates.append(
            {
                "codigo_usuario": user,
                "nome": dep["nome"],
                "cpf": cpf,
                "email": email,
                "telefone": telefone,
                "parentesco": dep["parentesco"],
                "sexo": dep["sexo"],
                "numero_dependente": dep["numero_dependente"],
                "data_nascimento": dep["data_nascimento"].isoformat() if dep["data_nascimento"] else None,
                "cpf_real": cpf_real,
            }
        )

    return tit_updates, dep_updates


def generate_sql(tit_updates: list[dict], dep_updates: list[dict]) -> str:
    lines = [
        "-- Migration 026: Sincroniza CPF, e-mail e telefone (titulares + dependentes) a partir da planilha",
        "-- Fonte: Relatorio_Associados_e_Dependentes.xlsx",
        "-- Titulares: bloco de contato na aba Titulares (cols 9-15) por matrícula",
        "-- Dependentes: bloco inline na aba Dependentes (cols 13-17) quando existir; senão CPF da coluna Senha",
        "",
        "BEGIN;",
        "",
    ]

    t_cpf = sum(1 for u in tit_updates if u["cpf_real"])
    t_email = sum(1 for u in tit_updates if u["email"])
    t_tel = sum(1 for u in tit_updates if u["telefone"])
    d_cpf = sum(1 for u in dep_updates if u["cpf_real"])
    d_email = sum(1 for u in dep_updates if u["email"])
    d_tel = sum(1 for u in dep_updates if u["telefone"])

    lines.append(
        f"-- Titulares: {len(tit_updates)} | CPF real: {t_cpf} | E-mail: {t_email} | Telefone: {t_tel}"
    )
    lines.append(
        f"-- Dependentes: {len(dep_updates)} | CPF real: {d_cpf} | E-mail: {d_email} | Telefone: {d_tel}"
    )
    lines.append("")

    for u in tit_updates:
        sets = [f"nome = {sql_str(u['nome'])}"]
        if u["cpf"]:
            sets.append(f"cpf = {sql_str(u['cpf'])}")
        sets.append(f"email = {sql_str(u['email'])}")
        sets.append(f"telefone = {sql_str(u['telefone'])}")
        if u["data_admissao"]:
            sets.append(f"data_admissao = {sql_str(u['data_admissao'])}")

        lines.append(
            f"UPDATE usuarios SET {', '.join(sets)} "
            f"WHERE codigo_usuario = {sql_str(u['codigo_usuario'])} AND categoria_socio = 'titular';"
        )

    lines.append("")

    for u in dep_updates:
        sets = [f"nome = {sql_str(u['nome'])}"]
        if u["cpf"]:
            sets.append(f"cpf = {sql_str(u['cpf'])}")
        sets.append(f"email = {sql_str(u['email'])}")
        sets.append(f"telefone = {sql_str(u['telefone'])}")
        if u["data_nascimento"]:
            sets.append(f"data_nascimento = {sql_str(u['data_nascimento'])}")
        if u["parentesco"]:
            sets.append(f"parentesco = {sql_str(u['parentesco'])}")
        if u["sexo"]:
            sets.append(f"sexo = {sql_str(u['sexo'])}")
        if u["numero_dependente"] is not None:
            sets.append(f"numero_dependente = {u['numero_dependente']}")

        lines.append(
            f"UPDATE usuarios SET {', '.join(sets)} "
            f"WHERE codigo_usuario = {sql_str(u['codigo_usuario'])} AND categoria_socio = 'dependente';"
        )

    lines.extend(["", "COMMIT;", ""])
    return "\n".join(lines)


def main() -> None:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx.exists():
        raise SystemExit(f"Planilha não encontrada: {xlsx}")

    tit_updates, dep_updates = build_updates(xlsx)
    sql = generate_sql(tit_updates, dep_updates)
    OUT_SQL.write_text(sql, encoding="utf-8")

    print(f"Planilha: {xlsx}")
    print(f"Titulares com UPDATE: {len(tit_updates)}")
    print(f"  CPF real: {sum(1 for u in tit_updates if u['cpf_real'])}")
    print(f"  E-mail: {sum(1 for u in tit_updates if u['email'])}")
    print(f"  Telefone: {sum(1 for u in tit_updates if u['telefone'])}")
    print(f"Dependentes com UPDATE: {len(dep_updates)}")
    print(f"  CPF real: {sum(1 for u in dep_updates if u['cpf_real'])}")
    print(f"  E-mail: {sum(1 for u in dep_updates if u['email'])}")
    print(f"  Telefone: {sum(1 for u in dep_updates if u['telefone'])}")
    print(f"SQL: {OUT_SQL}")


if __name__ == "__main__":
    main()
